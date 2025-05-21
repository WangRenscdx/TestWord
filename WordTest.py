import random
import tkinter as tk
from datetime import datetime, timedelta
from tkinter import messagebox, ttk, filedialog
import sqlite3
from pathlib import Path
import json
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

plt.rcParams["font.sans-serif"] = ["SimHei"]
plt.rcParams["axes.unicode_minus"] = False


class VocabularyTestApp:
    def __init__(self, root):
        self.root = root
        self.root.title("智能单词测验系统")
        self.root.state('zoomed')

        # 初始化变量
        self.mode = tk.StringVar(value="word_to_meaning")
        self.db_path = Path("vocabulary.db")
        self.words = []
        self.history_records = []
        self.current_words_display = []
        self.test_words = []
        self.current_question = 0
        self.correct_answers = 0
        self.time_left = timedelta(minutes=20)
        self.timer_running = False
        self.incorrect_words = []  # 存储本次测试中的错误单词

        # 初始化界面
        self.setup_styles()
        self.create_widgets()
        self.setup_database()
        self.load_data()
        self.show_home()

    def setup_styles(self):
        """配置界面样式"""
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TButton', font=('Microsoft YaHei', 11))
        style.configure('Nav.TButton', background='#F0F0F0', padding=5)
        style.map('Nav.TButton', background=[('active', '#0078D7')])
        style.configure('Accent.TButton', background='#0078D7', foreground='white')
        style.configure('Option.TButton',
                        padding=8,  # 增加内边距，提高可点击区域
                        anchor='w',  # 文本左对齐
                        justify='left')  # 多行文本左对齐

    def create_widgets(self):
        """创建主界面组件"""
        # 导航栏框架
        nav_frame = ttk.Frame(self.root, width=150)
        nav_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)

        # 导航按钮
        nav_buttons = [
            ("🏠 首页", self.show_home),
            ("📝 开始测验", self.show_test),
            ("📖 生词本", self.show_vocabulary),
            ("📊 统计", self.show_statistics),
            ("📤 导出", self.show_export),
            ("📥 导入", self.import_excel)
        ]
        for text, cmd in nav_buttons:
            ttk.Button(nav_frame, text=text, command=cmd, style='Nav.TButton').pack(fill=tk.X, pady=2)

        # 主内容区域
        self.main_content = ttk.Frame(self.root)
        self.main_content.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    def setup_database(self):
        """初始化数据库结构"""
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("""CREATE TABLE IF NOT EXISTS words
                            (
                                id
                                INTEGER
                                PRIMARY
                                KEY,
                                word
                                TEXT
                                NOT
                                NULL
                                UNIQUE,
                                pos
                                TEXT,
                                meaning
                                TEXT
                                NOT
                                NULL
                            )""")
            conn.execute("""CREATE TABLE IF NOT EXISTS history
                            (
                                id
                                INTEGER
                                PRIMARY
                                KEY,
                                test_date
                                TEXT
                                NOT
                                NULL,
                                accuracy
                                REAL
                                NOT
                                NULL,
                                duration
                                TEXT
                                NOT
                                NULL,
                                total_questions
                                INTEGER
                                NOT
                                NULL,
                                incorrect_words
                                TEXT -- 存储JSON格式的错误单词列表
                            )""")

    def load_data(self):
        """从数据库加载数据"""
        with sqlite3.connect(self.db_path) as conn:
            self.words = conn.execute("SELECT word, pos, meaning FROM words").fetchall()
            self.history_records = conn.execute(
                """SELECT test_date, accuracy, duration, total_questions, incorrect_words
                   FROM history
                                                   ORDER BY test_date DESC""").fetchall()
        self.current_words_display = self.words.copy()

    def clear_content(self):
        """清空内容区域"""
        for widget in self.main_content.winfo_children():
            widget.destroy()

    # 首页模块 ----------------------------------------------------------
    def show_home(self):
        """显示首页"""
        self.clear_content()
        ttk.Label(self.main_content, text="欢迎使用智能单词测验系统",
                  font=('Microsoft YaHei', 18)).pack(pady=50)
        ttk.Label(self.main_content,
                  text="\n\n功能导航\n\n📚 多种测验模式\n📈 学习进度追踪\n📤 数据导入导出",
                  font=('Microsoft YaHei', 14)).pack(expand=True)

    # 测验模块 ----------------------------------------------------------
    def show_test(self):
        """显示测验界面"""
        self.clear_content()

        # 模式选择
        mode_frame = ttk.LabelFrame(self.main_content, text="测验模式")
        mode_frame.pack(fill=tk.X, pady=5)
        for text, value in [("单词→释义", "word_to_meaning"),
                            ("释义→单词", "meaning_to_word"),
                            ("翻译填空", "translation_fill")]:
            ttk.Radiobutton(mode_frame, text=text, variable=self.mode,
                            value=value).pack(side=tk.LEFT, padx=10)

        # 开始按钮
        ttk.Button(self.main_content, text="开始测验",
                   command=self.start_test, style='Accent.TButton').pack(pady=20)

    def start_test(self):
        """开始测验"""
        if not self.words:
            messagebox.showwarning("提示", "单词本为空，请先导入数据！")
            return

        self.test_words = random.sample(self.words, min(20, len(self.words)))
        self.current_question = 0
        self.correct_answers = 0
        self.time_left = timedelta(minutes=20)
        self.timer_running = True
        self.incorrect_words = []  # 清空上次的错误记录

        self.clear_content()
        self.create_test_ui()
        self.update_timer()
        self.show_question()

    def create_test_ui(self):
        """创建测验界面组件"""
        # 控制面板
        control_frame = ttk.Frame(self.main_content)
        control_frame.pack(fill=tk.X, pady=5)

        # 进度显示
        self.progress_label = ttk.Label(control_frame, font=('Microsoft YaHei', 12))
        self.progress_label.pack(side=tk.LEFT)

        # 计时器
        self.timer_label = ttk.Label(control_frame, font=('Microsoft YaHei', 14))
        self.timer_label.pack(side=tk.RIGHT)

        # 题目区域
        self.question_frame = ttk.LabelFrame(self.main_content, text="题目")
        self.question_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.question_label = ttk.Label(self.question_frame, font=('Microsoft YaHei', 16))
        self.question_label.pack(pady=20)

        # 选项区域
        self.options_frame = ttk.Frame(self.main_content)
        self.options_frame.pack(fill=tk.BOTH, expand=True)

        # 结果区域
        self.result_label = ttk.Label(self.main_content, font=('Microsoft YaHei', 14))
        self.result_label.pack()

    def show_question(self):
        """显示题目"""
        if self.current_question >= len(self.test_words):
            self.end_test()
            return

        word, pos, meaning = self.test_words[self.current_question]
        mode = self.mode.get()

        # 更新进度
        self.progress_label.config(
            text=f"进度：{self.current_question + 1}/{len(self.test_words)}")

        # 清除旧内容
        for widget in self.options_frame.winfo_children():
            widget.destroy()

        # 处理不同模式
        if mode == "translation_fill":
            self.show_translation_fill(word, pos, meaning)
        else:
            question_text = {
                "word_to_meaning": f"单词：{word}\n词性：{pos}",
                "meaning_to_word": f"释义：{meaning}"
            }[mode]
            self.question_label.config(text=question_text)

            # 生成6个选项
            options = self.generate_options(
                correct=meaning if mode == "word_to_meaning" else word,
                count=6,
                field="meaning" if mode == "word_to_meaning" else "word"
            )

            for i, opt in enumerate(options):
                # 创建选项文本，添加序号前缀
                option_text = f"{chr(65 + i)}. {opt}"

                # 直接创建按钮，不使用嵌套Frame
                btn = ttk.Button(self.options_frame,
                                 text=option_text,
                                 style='Option.TButton',
                                 command=lambda o=opt: self.check_answer(o, meaning if mode == "word_to_meaning" else word))
                btn.pack(fill=tk.X, padx=5, pady=5)  # 使用pack布局，简化结构

    def show_translation_fill(self, word, pos, meaning):
        """显示翻译填空题"""
        masked_word = self.mask_word(word, 3, 5)
        self.question_label.config(text=f"释义：{meaning}\n\n请补全单词：{masked_word}")
        self.entry = ttk.Entry(self.options_frame, font=('Microsoft YaHei', 14))
        self.entry.pack(pady=20)
        self.entry.bind("<Return>", lambda e: self.check_fill_answer(word))

    def mask_word(self, word, min_mask=3, max_mask=5):
        """生成填空单词"""
        length = len(word)
        mask_num = min(max(random.randint(min_mask, max_mask), 1), length - 1)
        positions = random.sample(range(length), mask_num)
        return "".join(["_" if i in positions else c for i, c in enumerate(word)])

    def generate_options(self, correct, count, field):
        """生成选项"""
        options = [correct]
        while len(options) < count:
            random_item = random.choice(self.words)[0 if field == "word" else 2]
            if random_item not in options:
                options.append(random_item)
        random.shuffle(options)
        return options

    def check_answer(self, selected, correct):
        """检查选项答案"""
        if selected == correct:
            self.correct_answers += 1
            self.result_label.config(text="✓ 正确！", foreground="green")
        else:
            self.result_label.config(text=f"✗ 错误！正确答案：{correct}", foreground="red")
            # 记录错误单词
            current_word = self.test_words[self.current_question][0]
            self.incorrect_words.append({
                "word": current_word,
                "correct_meaning": correct,
                "user_answer": selected,
                "test_date": datetime.now().strftime("%Y-%m-%d %H:%M")
            })
        self.current_question += 1
        self.root.after(1000, self.show_question)

    def check_fill_answer(self, correct):
        """检查填空题答案"""
        answer = self.entry.get().strip()
        if answer.lower() == correct.lower():
            self.correct_answers += 1
            self.result_label.config(text="✓ 正确！", foreground="green")
        else:
            self.result_label.config(text=f"✗ 错误！正确答案：{correct}", foreground="red")
            # 记录错误单词
            current_word = self.test_words[self.current_question][0]
            self.incorrect_words.append({
                "word": current_word,
                "correct_meaning": correct,
                "user_answer": answer,
                "test_date": datetime.now().strftime("%Y-%m-%d %H:%M")
            })
        self.current_question += 1
        self.root.after(1000, self.show_question)

    def update_timer(self):
        """更新计时器"""
        if not self.timer_running:
            return

        if self.time_left.total_seconds() <= 0:
            self.end_test()
            return
        if not self.timer_label.winfo_exists():
            self.timer_running = False
            return
        self.time_left -= timedelta(seconds=1)
        mins, secs = divmod(int(self.time_left.total_seconds()), 60)
        self.timer_label.config(text=f"{mins:02}:{secs:02}")
        self.root.after(1000, self.update_timer)
        # 使用try-except块捕获可能的错误
        try:
            self.timer_label.config(text=f"{mins:02}:{secs:02}")
        except:
            self.timer_running = False

        # 仅在标签存在且计时器运行时继续计时
        if self.timer_label.winfo_exists() and self.timer_running:
            self.root.after(1000, self.update_timer)

    def end_test(self):
        """结束测试"""
        self.timer_running = False
        total = len(self.test_words)
        accuracy = round(self.correct_answers / total * 100, 1) if total else 0
        duration = str(datetime.now() - (datetime.now() - self.time_left)).split(".")[0]

        # 保存历史记录
        incorrect_data = json.dumps(self.incorrect_words) if self.incorrect_words else None
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("""INSERT INTO history (test_date, accuracy, duration, total_questions, incorrect_words)
                            VALUES (?, ?, ?, ?, ?)""",
                         (datetime.now().strftime("%Y-%m-%d %H:%M"),
                          accuracy,
                          duration,
                          total,
                          incorrect_data))

        messagebox.showinfo("测试完成",
                            f"正确率：{accuracy}%\n用时：{duration}\n正确题数：{self.correct_answers}/{total}")
        self.load_data()  # 重新加载数据以更新历史记录
        self.show_statistics()  # 直接跳转到统计页面

    # 生词本模块 --------------------------------------------------------
    def show_vocabulary(self):
        """显示生词本"""
        self.clear_content()

        # 搜索框
        search_frame = ttk.Frame(self.main_content)
        search_frame.pack(fill=tk.X, pady=5)
        self.search_var = tk.StringVar()
        ttk.Entry(search_frame, textvariable=self.search_var, width=30).pack(side=tk.LEFT, padx=5)
        ttk.Button(search_frame, text="搜索", command=self.search_words).pack(side=tk.LEFT)

        # 单词表格
        columns = ("单词", "词性", "释义")
        self.tree = ttk.Treeview(self.main_content, columns=columns, show="headings")
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 分页控制
        self.current_page = 1
        self.words_per_page = 20
        pagination = ttk.Frame(self.main_content)
        pagination.pack(pady=5)
        ttk.Button(pagination, text="上一页", command=lambda: self.change_page(-1)).pack(side=tk.LEFT)
        self.page_label = ttk.Label(pagination, text="第1页/共1页")
        self.page_label.pack(side=tk.LEFT, padx=10)
        ttk.Button(pagination, text="下一页", command=lambda: self.change_page(1)).pack(side=tk.LEFT)

        # 操作按钮
        btn_frame = ttk.Frame(self.main_content)
        btn_frame.pack(pady=5)
        ttk.Button(btn_frame, text="添加单词", command=self.show_add_dialog).pack(side=tk.LEFT)
        ttk.Button(btn_frame, text="删除选中", command=self.delete_word).pack(side=tk.LEFT, padx=10)

        self.load_vocab_table()

    def load_vocab_table(self):
        """加载单词表格"""
        for item in self.tree.get_children():
            self.tree.delete(item)

        start = (self.current_page - 1) * self.words_per_page
        end = start + self.words_per_page
        for word, pos, meaning in self.words[start:end]:
            self.tree.insert("", "end", values=(word, pos, meaning))

        total_pages = (len(self.words) - 1) // self.words_per_page + 1
        self.page_label.config(text=f"第{self.current_page}页/共{total_pages}页")

    def search_words(self):
        """搜索单词"""
        keyword = self.search_var.get().strip().lower()
        if not keyword:
            self.load_vocab_table()
            return

        results = []
        for word, pos, meaning in self.words:
            if keyword in word.lower() or keyword in meaning.lower():
                results.append((word, pos, meaning))

        for item in self.tree.get_children():
            self.tree.delete(item)
        for item in results:
            self.tree.insert("", "end", values=item)

    def change_page(self, direction):
        """分页控制"""
        self.current_page += direction
        total_pages = (len(self.words) - 1) // self.words_per_page + 1
        if self.current_page < 1:
            self.current_page = 1
        elif self.current_page > total_pages:
            self.current_page = total_pages
        self.load_vocab_table()

    def show_add_dialog(self):
        """显示添加对话框"""
        dialog = tk.Toplevel()
        dialog.title("添加单词")

        ttk.Label(dialog, text="单词：").grid(row=0, column=0, padx=5, pady=5)
        ttk.Label(dialog, text="词性：").grid(row=1, column=0, padx=5, pady=5)
        ttk.Label(dialog, text="释义：").grid(row=2, column=0, padx=5, pady=5)

        entries = {
            'word': ttk.Entry(dialog),
            'pos': ttk.Entry(dialog),
            'meaning': ttk.Entry(dialog)
        }
        for i, (key, entry) in enumerate(entries.items()):
            entry.grid(row=i, column=1, padx=5, pady=5)

        ttk.Button(dialog, text="提交", command=lambda: self.add_word(entries)).grid(row=3, columnspan=2, pady=10)

    def add_word(self, entries):
        """添加单词"""
        word = entries['word'].get().strip()
        pos = entries['pos'].get().strip()
        meaning = entries['meaning'].get().strip()

        if not word or not meaning:
            messagebox.showwarning("错误", "单词和释义不能为空！")
            return

        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.execute("INSERT INTO words (word, pos, meaning) VALUES (?, ?, ?)",
                             (word, pos, meaning))
            self.load_data()
            self.load_vocab_table()
            messagebox.showinfo("成功", "单词添加成功！")
        except sqlite3.IntegrityError:
            messagebox.showerror("错误", "该单词已存在！")

    def delete_word(self):
        """删除单词"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要删除的单词")
            return

        word = self.tree.item(selected[0], 'values')[0]
        if messagebox.askyesno("确认", f"确定要删除 {word} 吗？"):
            with sqlite3.connect(self.db_path) as conn:
                conn.execute("DELETE FROM words WHERE word = ?", (word,))
            self.load_data()
            self.load_vocab_table()

    # 统计模块 ----------------------------------------------------------
    def show_statistics(self):
        """显示统计信息（包含错误单词列表）"""
        self.clear_content()

        # 创建主框架（分上下两部分：图表 + 错误单词表）
        main_frame = ttk.Frame(self.main_content)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 上半部分：正确率趋势图
        chart_frame = ttk.Frame(main_frame, height=200)
        chart_frame.pack(fill=tk.X, padx=5, pady=5)

        # 下半部分：错误单词表
        error_frame = ttk.Frame(main_frame)
        error_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # ---------------------- 绘制图表 ----------------------
        fig = plt.Figure(figsize=(8, 3), dpi=100)
        ax = fig.add_subplot(111)
        plt.rcParams["font.family"] = ["SimHei"]
        plt.rcParams["axes.unicode_minus"] = False

        if self.history_records:
            recent_records = self.history_records[:10][::-1]
            dates = [rec[0][5:10] for rec in recent_records]
            accuracies = [rec[1] for rec in recent_records]
            ax.plot(dates, accuracies, marker='o', color='#4a86e8')
            ax.set_title("最近10次测试正确率趋势", fontsize=12)
            ax.set_xlabel("日期", fontsize=10)
            ax.set_ylabel("正确率 (%)", fontsize=10)
            ax.grid(True, linestyle='--', alpha=0.7)
        else:
            ax.text(0.5, 0.5, "暂无测试数据", ha='center', va='center')
            ax.axis('off')

        fig.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.X)

        # ---------------------- 显示错误单词表 ----------------------
        # 获取最近的错误单词记录（最多10条）
        recent_errors = self.get_recent_incorrect_words()

        if not recent_errors:
            ttk.Label(error_frame, text="暂无错误单词记录", font=('Microsoft YaHei', 12)).pack(pady=20)
            return

        # 创建表格
        columns = ("测试日期", "单词", "正确释义", "你的答案")
        self.error_tree = ttk.Treeview(error_frame, columns=columns, show="headings", height=5)
        for col in columns:
            self.error_tree.heading(col, text=col)
            self.error_tree.column(col, width=120 if col == "测试日期" else 180)  # 调整列宽
        self.error_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 填充数据
        for error in recent_errors:
            self.error_tree.insert("", "end", values=(
                error["test_date"][:16],  # 缩短日期显示（只显示到分钟）
                error["word"],
                error["correct_meaning"],
                error["user_answer"]
            ))

        # 添加滚动条
        scrollbar = ttk.Scrollbar(error_frame, orient="vertical", command=self.error_tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.error_tree.configure(yscrollcommand=scrollbar.set)

        # 添加标题
        ttk.Label(error_frame, text="最近错误单词（最多显示10条）", font=('Microsoft YaHei', 12, 'bold')).pack(pady=5, anchor='w')

    def get_recent_incorrect_words(self):
        """从数据库获取最近的错误单词记录（最多10条）"""
        recent_errors = []
        with sqlite3.connect(self.db_path) as conn:
            # 按时间倒序查询最近10条包含错误单词的记录
            records = conn.execute("""
                                   SELECT test_date, incorrect_words
                                   FROM history
                                   WHERE incorrect_words IS NOT NULL
                                   ORDER BY test_date DESC
                                       LIMIT 10
                                   """).fetchall()
        for date_str, json_data in records:
            try:
                errors = json.loads(json_data)  # 解析JSON数据
                recent_errors.extend(errors)
            except json.JSONDecodeError:
                continue
        return recent_errors[:10]  # 最多显示10条错误记录

    # 导出模块 ----------------------------------------------------------
    def show_export(self):
        """显示导出界面"""
        self.clear_content()

        export_frame = ttk.LabelFrame(self.main_content, text="导出选项")
        export_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        formats = [
            ("Excel 文件 (.xlsx)", self.export_excel),
            ("PDF 文件 (.pdf)", self.export_pdf),
            ("文本文件 (.txt)", self.export_text)
        ]

        for text, cmd in formats:
            ttk.Button(export_frame, text=text, command=cmd).pack(pady=5)

    def export_excel(self):
        """导出到Excel"""
        try:
            from openpyxl import Workbook
        except ImportError:
            messagebox.showerror("错误", "请先安装openpyxl库：pip install openpyxl")
            return

        path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if not path: return

        try:
            wb = Workbook()
            ws = wb.active
            ws.append(["单词", "词性", "释义"])
            for word, pos, meaning in self.words:
                ws.append([word, pos, meaning])
            wb.save(path)
            messagebox.showinfo("成功", f"已导出到 {path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{str(e)}")

    def export_pdf(self):
        """导出到PDF"""
        try:
            from fpdf import FPDF
        except ImportError:
            messagebox.showerror("错误", "请先安装fpdf库：pip install fpdf")
            return

        path = filedialog.asksaveasfilename(defaultextension=".pdf")
        if not path: return

        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.add_font('SimHei', '', 'simhei.ttf', uni=True)
            pdf.set_font('SimHei', '', 12)

            col_widths = [40, 20, 130]
            pdf.cell(col_widths[0], 10, "单词", border=1)
            pdf.cell(col_widths[1], 10, "词性", border=1)
            pdf.cell(col_widths[2], 10, "释义", border=1)
            pdf.ln()

            for word, pos, meaning in self.words:
                pdf.cell(col_widths[0], 10, word, border=1)
                pdf.cell(col_widths[1], 10, pos or "", border=1)
                pdf.cell(col_widths[2], 10, meaning, border=1)
                pdf.ln()

            pdf.output(path)
            messagebox.showinfo("成功", f"已导出到 {path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{str(e)}")

    def export_text(self):
        """导出到文本文件"""
        path = filedialog.asksaveasfilename(defaultextension=".txt")
        if not path: return

        try:
            with open(path, 'w', encoding='utf-8') as f:
                f.write("单词\t词性\t释义\n")
                f.write("-" * 50 + "\n")
                for word, pos, meaning in self.words:
                    f.write(f"{word}\t{pos}\t{meaning}\n")
            messagebox.showinfo("成功", f"已导出到 {path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{str(e)}")

    # 导入模块 ----------------------------------------------------------
    def import_excel(self):
        """导入Excel数据"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            messagebox.showerror("错误", "请先安装openpyxl库：pip install openpyxl")
            return

        path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx")])
        if not path: return

        try:
            wb = load_workbook(path)
            ws = wb.active
            new_words = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    new_words.append((
                        str(row[0]).strip(),
                        str(row[1]).strip() if len(row) > 1 else "",
                        str(row[2]).strip() if len(row) > 2 else ""
                    ))

            with sqlite3.connect(self.db_path) as conn:
                conn.executemany("""INSERT OR IGNORE INTO words (word, pos, meaning)
                                  VALUES (?, ?, ?)""", new_words)

            self.load_data()
            messagebox.showinfo("成功", f"成功导入 {len(new_words)} 条记录")
        except Exception as e:
            messagebox.showerror("错误", f"导入失败：{str(e)}")


if __name__ == "__main__":
    root = tk.Tk()
    app = VocabularyTestApp(root)
    root.mainloop()